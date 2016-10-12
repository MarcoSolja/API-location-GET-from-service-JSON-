__author__ = 'msawyer'

"""
ZipCodeAPIResult Project
by: Mark Sawyer
Intent: Get a list of failed zipcode entry running -> https://api2.iheart.com/api/v2/content/liveStations?
Took zipcode list from http://federalgovernmentzipcodes.us/
"""

# import json
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

urlDomain = 'https://www.iheart.com'
locService = '/api/v2/content/liveStations?offset=0&allMarkets=false&limit=10'

def main():

    zips, zipCodeTypes, cities, states, lats, longs = ([] for i in range(6))

    # 1.0 Create the workbook to get data from
    wb = Workbook()
    wb = load_workbook('/Users/msawyer/Downloads/free-zipcode-database-Primary.xlsx') # Took zipcode list from http://federalgovernmentzipcodes.us

    # 1.1 Get the zipcode, type, city and state and add them to an array
    # 1.2 grab the active worksheet
    ws = wb.active

    for row in ws.iter_rows():
        zips += [row[0].value]
        zipCodeTypes += [row[1].value]
        cities += [row[2].value]
        states += [row[3].value]
        lats += [row[5].value]
        longs += [row[6].value]

    # 1.3 close workbook
    wb.save('/Users/msawyer/Downloads/TestWB.xlsx')

    # 2 Create another clean wb to enter results to
    wb = Workbook()
    ws = wb.active

    # 2.1 Go through each zip and lat/long and deliver the payload

    writezipstoxls(zips, zipCodeTypes, cities, states, lats, longs, wb, ws)


def writezipstoxls(zips, zipCodeTypes, cities, states, lats, longs, wb, ws):
    icount = 1
    for zip in zips:
        if icount != 1:
            completeZipCall = urlDomain + locService + '&zipCode=' + str(zips[icount - 1])
            completeLatLongCall = urlDomain + locService + '&lat=' + str(lats[icount - 1]) + '&lng=' + str(
                longs[icount - 1])

            # 2.2 Start populating cells with known data
            ws.cell('A' + str(icount)).value = str(zips[icount - 1])
            ws.cell('B' + str(icount)).value = str(zipCodeTypes[icount - 1])
            ws.cell('C' + str(icount)).value = str(cities[icount - 1])
            ws.cell('D' + str(icount)).value = str(states[icount - 1])
            ws.cell('E' + str(icount)).value = str(lats[icount - 1])
            ws.cell('F' + str(icount)).value = str(longs[icount - 1])

            # 2.3 Start populating cells with location results
            responseZipCall = requests.get(completeZipCall)
            responseLatLongCall = requests.get(completeLatLongCall)

            # 2.4 Parse json results for location Pass / Failure
            jsonZipCall = responseZipCall.json()
            jsonLatLongCall = responseLatLongCall.json()

            ws.cell('G' + str(icount)).value = completeZipCall
            ws.cell('I' + str(icount)).value = completeLatLongCall

            ws.cell('K' + str(icount)).value = str(
                jsonZipCall)  # This was a great idea but really no human eye will read
            # this. Detection needs to be smart.
            ws.cell('L' + str(icount)).value = str(jsonLatLongCall)  # This was a great idea but really no human eye
            # will read this. Detection needs to be smart.
        else:
            # Column name assignment
            ws.cell('A1').value = 'ZipCode'
            ws.cell('B1').value = 'Zip Code Type'
            ws.cell('C1').value = 'City'
            ws.cell('D1').value = 'State'
            ws.cell('E1').value = 'Latitude'
            ws.cell('F1').value = 'Longitude'
            ws.cell('G1').value = 'ZipCode entry API'
            ws.cell('H1').value = 'ZipCode entry PASS/FAIL'
            ws.cell('I1').value = 'Lat/Long entry API'
            ws.cell('J1').value = 'Lat/Long entry PASS/FAIL'
            ws.cell('K1').value = 'JSON Reply - Zipcode'
            ws.cell('L1').value = 'JSON Reply - Lat/Long'

        icount += 1
    wb.save("/Users/msawyer/Downloads/AMPTestResult1.xlsx")

main()